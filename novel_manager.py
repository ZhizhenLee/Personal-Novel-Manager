# -*- coding: utf-8 -*-
"""通天录 - 小说编写管理器 (Personal Novel Manager)

单文件 Windows 桌面程序：
  - 每个小说一个主目录，目录名即小说名
  - 内含 草稿/ 封面/ 角色库/ 正文/ 世界观.md 章节标题列表.md 章节简介汇总.md _备份/ 及其他文件
  - 支持 新建/编辑/查看 草稿与正文、角色库、世界观、章节标题与简介
  - 支持配置小说根目录路径（持久化到用户主目录）
  - 支持 RAR 导入覆盖整个小说目录（解压软件路径可配置并记忆）
  - 支持将整个小说导出为 RAR 文件
  - 封面图片上传与显示
"""

import os
import sys
import re
import json
import shutil
import tempfile
import datetime
import subprocess

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog, scrolledtext
from tkinter import font as tkfont

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except Exception:
    HAS_PIL = False

APP_NAME = "通天录"
APP_VERSION = "1.0.0"

CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".tongtianlu_config.json")

# 小说主目录内的标准结构
SUBDIR_DRAFTS = "草稿"
SUBDIR_COVER = "封面"
SUBDIR_ROLES = "角色库"
SUBDIR_BODY = "正文"
SUBDIR_BACKUP = "_备份"
SUBDIR_OTHER = "__other__"

FILE_WORLDVIEW = "世界观.md"
FILE_CHAPTER_TITLES = "章节标题列表.md"
FILE_CHAPTER_SUMMARY = "章节简介汇总.md"

# 文本文件编码
TEXT_ENCODINGS = ["utf-8", "gbk", "utf-8-sig", "big5"]


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def read_text(path):
    """尝试多种编码读取文本文件，返回字符串。失败返回空字符串。"""
    if not os.path.exists(path):
        return ""
    raw = open(path, "rb").read()
    for enc in TEXT_ENCODINGS:
        try:
            return raw.decode(enc)
        except Exception:
            continue
    try:
        return raw.decode("utf-8", errors="replace")
    except Exception:
        return ""


def write_text(path, content):
    """以 UTF-8 写入文本文件，确保目录存在。"""
    ensure_dir(os.path.dirname(path))
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(content)


def write_text_smart(path, content):
    """写入，保持原编码（若原文件存在且识别为 gbk 则用 gbk，否则 utf-8）。"""
    ensure_dir(os.path.dirname(path))
    if os.path.exists(path):
        raw = open(path, "rb").read()
        try:
            raw.decode("utf-8")
            encoding = "utf-8"
        except Exception:
            # 尝试 gbk
            try:
                raw.decode("gbk")
                if "\x00" in raw.decode("utf-8", errors="ignore"):
                    encoding = "utf-8"
                else:
                    encoding = "gbk"
            except Exception:
                encoding = "utf-8"
    else:
        encoding = "utf-8"
    with open(path, "w", encoding=encoding, newline="") as f:
        f.write(content)


def ensure_dir(path):
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def timestamp_str():
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S")


def draft_default_name(ext=".md"):
    return datetime.datetime.now().strftime("%Y%m%d%H%M") + ext


def is_image_file(name):
    return name.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".ico"))


def _read_xlsx_rows(path, max_rows=500, max_cols=60):
    """读取 xlsx 第一张工作表的二维数据。

    优先使用 openpyxl（若可用，兼容性更好）；失败则回退到标准库极简解析。
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([("" if v is None else str(v)) for v in row[:max_cols]])
        wb.close()
        return rows
    except Exception:
        pass
    # 回退：标准库极简解析
    import zipfile
    import xml.etree.ElementTree as ET
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    z = zipfile.ZipFile(path)
    shared = []
    if "xl/sharedStrings.xml" in z.namelist():
        root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for si in root.findall(ns + "si"):
            shared.append("".join(t.text or "" for t in si.iter(ns + "t")))
    sheets = sorted([n for n in z.namelist()
                    if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")])
    if not sheets:
        return []
    root = ET.fromstring(z.read(sheets[0]))
    rows = []
    for row in root.iter(ns + "row"):
        cells = []
        for c in row.findall(ns + "c"):
            t = c.get("t")
            v = c.find(ns + "v")
            isn = c.find(ns + "is")
            val = ""
            if t == "s" and v is not None:
                try:
                    val = shared[int(v.text)]
                except Exception:
                    val = ""
            elif isn is not None:
                val = "".join(tt.text or "" for tt in isn.iter(ns + "t"))
            elif v is not None:
                val = v.text or ""
            cells.append(val)
        rows.append(cells)
        if len(rows) >= max_rows:
            break
    return [r[:max_cols] for r in rows]


def _xlsx_fill_color(cell):
    """返回单元格填充背景色（#rrggbb），无填充返回 None。"""
    try:
        fill = cell.fill
        if fill is None or getattr(fill, "patternType", None) != "solid":
            return None
        col = fill.fgColor
        if col is None or getattr(col, "type", None) != "rgb":
            return None
        rgb = col.rgb
        if not rgb or rgb in ("00000000", "FFFFFFFF"):
            return None
        if len(rgb) == 8:
            rgb = rgb[2:]
        return "#" + rgb
    except Exception:
        return None


def _xlsx_anchor(cell):
    """返回单元格水平对齐对应的 tk anchor。"""
    try:
        h = cell.alignment.horizontal
        if h == "center" or h == "centerContinuous":
            return "center"
        if h == "right":
            return "e"
        if h == "left" or h == "justify":
            return "w"
    except Exception:
        pass
    return "w"


def _xlsx_font_color(cell):
    """返回单元格字体颜色（#rrggbb），未显式设置则返回 None。"""
    try:
        f = cell.font
        col = getattr(f, "color", None)
        if col is None:
            return None
        if getattr(col, "type", None) == "rgb":
            rgb = col.rgb
            if rgb and rgb not in ("00000000",):
                if len(rgb) == 8:
                    rgb = rgb[2:]
                return "#" + rgb
    except Exception:
        pass
    return None


def _xlsx_border(cell):
    """返回 (是否有边框, 边框颜色)。仅当某条边具有边框样式时记为有边框。"""
    try:
        b = cell.border
        sides = [b.top, b.bottom, b.left, b.right]
        if not any(getattr(s, "style", None) not in (None, "none") for s in sides):
            return False, None
        color = None
        for s in sides:
            c = getattr(s, "color", None)
            if c is not None and getattr(c, "type", None) == "rgb" \
                    and c.rgb and c.rgb not in ("00000000",):
                rgb = c.rgb
                if len(rgb) == 8:
                    rgb = rgb[2:]
                color = "#" + rgb
                break
        return True, color
    except Exception:
        return False, None


def _cell_ref_to_rc(ref):
    """将单元格引用（如 'AB12'）转为 (row, col)，行列均从 1 开始。"""
    import re
    m = re.match(r"([A-Z]+)(\d+)$", ref)
    letters, row = m.group(1), int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    return row, col


def _xlsx_merged_ranges(path):
    """读取 xlsx 第一张工作表的合并单元格，返回 (merged, covered)。

    merged: {(ri,ci): (rowspan, colspan)}（0 基，左上角）
    covered: 被合并覆盖（非左上角）的坐标集合
    使用标准库解析 XML，避免 read_only 模式下 openpyxl 无法读取合并信息。
    """
    import zipfile
    import re
    import xml.etree.ElementTree as ET
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    z = zipfile.ZipFile(path)
    sheet_file = None
    try:
        if "xl/workbook.xml" in z.namelist():
            wb = ET.fromstring(z.read("xl/workbook.xml"))
            sheets = wb.find(ns + "sheets")
            if sheets is not None and len(sheets):
                rid = sheets[0].get(rns + "id")
                rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
                for rel in rels:
                    if rel.get("Id") == rid:
                        tgt = rel.get("Target")
                        sheet_file = tgt.lstrip("/") if tgt.startswith("/") \
                            else "xl/" + tgt.lstrip("/")
                        break
    except Exception:
        sheet_file = None
    if not sheet_file:
        cand = sorted(n for n in z.namelist()
                      if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
        sheet_file = cand[0] if cand else None
    if not sheet_file:
        return {}, set()
    root = ET.fromstring(z.read(sheet_file))
    merges = []
    mc = root.find(ns + "mergeCells")
    if mc is not None:
        for m in mc.findall(ns + "mergeCell"):
            ref = m.get("ref")
            if ref and ":" in ref:
                a, b = ref.split(":")
                merges.append((_cell_ref_to_rc(a), _cell_ref_to_rc(b)))
    merged, covered = {}, set()
    for (r0, c0), (r1, c1) in merges:
        rs = r1 - r0 + 1
        cs = c1 - c0 + 1
        merged[(r0 - 1, c0 - 1)] = (rs, cs)
        for dr in range(rs):
            for dc in range(cs):
                if dr or dc:
                    covered.add((r0 - 1 + dr, c0 - 1 + dc))
    return merged, covered


def _read_xlsx_with_styles(path, max_rows=300, max_cols=40):
    """读取 xlsx 第一张工作表，返回 (rows, total_rows, total_cols, merged, covered)。

    rows: [[{'v':值,'bg':颜色,'anchor':对齐}, ...], ...]
    merged: {(ri,ci): (rowspan, colspan)} 合并区域左上角
    covered: 被合并覆盖（非左上角）的单元格坐标集合
    优先 openpyxl 以读取样式与合并；失败则回退到无样式读取。
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        try:
            total_rows = int(ws.max_row or 0)
            total_cols = int(ws.max_column or 0)
        except Exception:
            total_rows = total_cols = 0
        merged = {}
        covered = set()
        try:
            merged, covered = _xlsx_merged_ranges(path)
        except Exception:
            merged, covered = {}, set()
        rows = []
        for i, row in enumerate(ws.iter_rows()):
            if i >= max_rows:
                break
            cells = []
            for c in row[:max_cols]:
                val = "" if c.value is None else str(c.value)
                bord, bcol = _xlsx_border(c)
                cells.append({"v": val, "bg": _xlsx_fill_color(c),
                              "fg": _xlsx_font_color(c),
                              "anchor": _xlsx_anchor(c),
                              "bordered": bord, "border_color": bcol})
            rows.append(cells)
        wb.close()
        return rows, total_rows, total_cols, merged, covered
    except Exception:
        pass
    data = _read_xlsx_rows(path, max_rows, max_cols)
    rows = [[{"v": ("" if v is None else str(v)), "bg": None, "fg": None,
              "anchor": "w", "bordered": False, "border_color": None} for v in r] for r in data]
    return rows, len(data), (max(len(r) for r in data) if data else 0), {}, set()


def load_config():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(cfg):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# 应用
# ---------------------------------------------------------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1200x760")
        self.minsize(900, 600)

        self.cfg = load_config()
        self.root_dir = self.cfg.get("root_dir", "")
        self._unrar_path = self.cfg.get("unrar_path", "")
        self.current_novel = None          # 当前打开的小说主目录

        # 字体
        base_font = tkfont.nametofont("TkDefaultFont")
        base_font.configure(family="Microsoft YaHei UI", size=10)

        self._build_style()
        self._build_menu()
        self._build_main_layout()

        self.show_novel_list()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---------------- 样式 ----------------
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except Exception:
            try:
                style.theme_use("clam")
            except Exception:
                pass

    # ---------------- 菜单 ----------------
    def _build_menu(self):
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="设置小说根目录...", command=self.set_root_dir)
        file_menu.add_command(label="新建小说...", command=self.new_novel)
        file_menu.add_separator()
        file_menu.add_command(label="导入RAR覆盖小说目录...", command=self.import_rar)
        file_menu.add_command(label="导出小说为RAR文件...", command=self.export_novel_dialog)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self._on_close)
        menubar.add_cascade(label=" 文件 ", menu=file_menu)

        tool_menu = tk.Menu(menubar, tearoff=0)
        tool_menu.add_command(label="刷新小说列表", command=self.refresh_novel_list)
        menubar.add_cascade(label=" 工具 ", menu=tool_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="关于", command=self.show_about)
        menubar.add_cascade(label=" 帮助 ", menu=help_menu)
        self.config(menu=menubar)

    # ---------------- 主布局(单容器，用于切换不同视图) ----------------
    def _build_main_layout(self):
        self.workspace = None
        self.current_view = None
        self._main_container = ttk.Frame(self)
        self._main_container.pack(fill="both", expand=True)

    def show_novel_list(self):
        """主界面：小说列表（封面图片+小说名）。"""
        if self.current_view is not None:
            self.current_view.destroy()
            self.current_view = None
        self.workspace = None
        self.current_view = NovelListView(self._main_container, self)
        self.current_view.pack(fill="both", expand=True)

    def show_workspace(self, novel_dir):
        """内容管理界面：左侧文件夹 + 中间文件列表 + 右侧正文。"""
        if self.current_view is not None:
            try:
                sv = getattr(self.current_view, "save_current", None)
                if sv:
                    sv()
            except Exception:
                pass
            self.current_view.destroy()
            self.current_view = None
        self.workspace = None
        self.current_novel = novel_dir
        self.workspace = NovelWorkspace(self._main_container, self, novel_dir)
        self.workspace.pack(fill="both", expand=True)
        self.current_view = self.workspace

    # ---------------- 空占位(无根目录时提示) ----------------
    def _show_empty_placeholder(self):
        pass

    # ================= 小说列表 =================
    def find_novel_names(self):
        """扫描根目录，返回小说主目录名列表。"""
        names = []
        if self.root_dir and os.path.isdir(self.root_dir):
            try:
                items = sorted(os.listdir(self.root_dir))
            except Exception:
                items = []
            for name in items:
                p = os.path.join(self.root_dir, name)
                if os.path.isdir(p) and not name.startswith("_") and name != "备份":
                    names.append(name)
        return names

    def refresh_novel_list(self):
        """刷新主界面的小说列表。"""
        self._novel_names = self.find_novel_names()
        if self.current_view is not None and isinstance(self.current_view, NovelListView):
            self.current_view.refresh()

    def get_selected_novel_path(self):
        sel = self.current_view.selected if isinstance(self.current_view, NovelListView) else None
        if not sel:
            return None
        return os.path.join(self.root_dir, sel)

    # ================= 封面上传/显示 =================
    def _load_cover_photo(self, novel_dir, size=96):
        if not HAS_PIL:
            return None
        cover_dir = os.path.join(novel_dir, SUBDIR_COVER)
        if not os.path.isdir(cover_dir):
            return None
        try:
            names = [n for n in os.listdir(cover_dir)
                     if os.path.isfile(os.path.join(cover_dir, n)) and is_image_file(n)]
            if not names:
                return None
            names.sort()
            path = os.path.join(cover_dir, names[0])
            img = Image.open(path)
            img.thumbnail((size, size), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    def upload_cover(self, novel_dir=None):
        novel_dir = novel_dir or self.current_novel
        if not novel_dir:
            return
        path = filedialog.askopenfilename(
            title="选择封面图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.gif *.bmp *.webp *.ico"),
                       ("所有文件", "*.*")],
        )
        if not path:
            return
        cover_dir = os.path.join(novel_dir, SUBDIR_COVER)
        ensure_dir(cover_dir)
        ext = os.path.splitext(path)[1].lower() or ".png"
        target = os.path.join(cover_dir, "cover" + ext)
        shutil.copy2(path, target)
        messagebox.showinfo(APP_NAME, "封面上传成功。")
        if (isinstance(self.current_view, NovelWorkspace)
                and self.current_view.novel_dir == self.current_novel):
            self.current_view.refresh_cover()

    # ================= 新建/打开/删除 小说 =================
    def new_novel(self):
        if not self.root_dir or not os.path.isdir(self.root_dir):
            if not self._prompt_root_dir():
                return
        name = simpledialog.askstring("新建小说", "请输入小说名称：", parent=self)
        if not name:
            return
        name = name.strip().strip("/\\")
        if not name:
            return
        novel_dir = os.path.join(self.root_dir, name)
        if os.path.exists(novel_dir):
            messagebox.showwarning(APP_NAME, "同名小说已存在。")
            return
        self._create_novel_structure(novel_dir)
        self._novel_names = self.find_novel_names()
        if self.current_view is not None and isinstance(self.current_view, NovelListView):
            self.current_view.refresh()
            self.current_view.select_by_name(name)
        self.show_workspace(novel_dir)

    def _create_novel_structure(self, novel_dir):
        ensure_dir(novel_dir)
        ensure_dir(os.path.join(novel_dir, SUBDIR_DRAFTS))
        ensure_dir(os.path.join(novel_dir, SUBDIR_COVER))
        ensure_dir(os.path.join(novel_dir, SUBDIR_ROLES))
        ensure_dir(os.path.join(novel_dir, SUBDIR_BODY))
        # 示例正文第一章
        body_001 = os.path.join(novel_dir, SUBDIR_BODY, "chapter_001.txt")
        write_text(body_001, "# 第一章\n\n（在此编写正文……）")
        # 模板
        write_text(os.path.join(novel_dir, FILE_WORLDVIEW),
                   "# 世界观\n\n（编写小说的世界观设定……）")
        write_text(os.path.join(novel_dir, FILE_CHAPTER_TITLES),
                   "# 章节标题列表\n\n- 第一章 ——（标题）\n")
        write_text(os.path.join(novel_dir, FILE_CHAPTER_SUMMARY),
                   "# 章节简介汇总\n\n- 第一章 ——（简介）\n")

    def open_selected_novel(self):
        path = self.get_selected_novel_path()
        if path:
            self.show_workspace(path)

    def delete_selected_novel(self):
        path = self.get_selected_novel_path()
        if not path:
            return
        if not messagebox.askyesno("删除小说",
                                   "确定要删除整个小说目录吗？此操作不可恢复！\n" + path):
            return
        if self.current_novel == path:
            self.current_novel = None
        try:
            shutil.rmtree(path)
        except Exception as e:
            messagebox.showerror(APP_NAME, "删除失败：%s" % e)
            return
        self._novel_names = self.find_novel_names()
        if self.current_view is not None and isinstance(self.current_view, NovelListView):
            self.current_view.refresh()

    def _select_novel(self, name):
        if self.current_view is not None and isinstance(self.current_view, NovelListView):
            self.current_view.select_by_name(name)

    # ================= 工作区 =================
    def open_novel(self, novel_dir):
        self.show_workspace(novel_dir)

    # ================= 设置根目录 =================
    def set_root_dir(self):
        if self._prompt_root_dir():
            self._novel_names = self.find_novel_names()
            if self.current_view is not None and isinstance(self.current_view, NovelListView):
                self.current_view.refresh()

    def _prompt_root_dir(self):
        path = filedialog.askdirectory(title="选择小说根目录")
        if not path:
            return False
        self.root_dir = path
        self.cfg["root_dir"] = path
        save_config(self.cfg)
        return True

    # ================= 导入 RAR（覆盖小说目录） =================
    def import_rar(self):
        if not self.root_dir or not os.path.isdir(self.root_dir):
            if not self._prompt_root_dir():
                return
        novels = self._novel_names
        if not novels:
            messagebox.showinfo(APP_NAME, "当前根目录下没有可替换的小说。")
            return
        src = filedialog.askopenfilename(
            title="选择要导入的 RAR 文件",
            filetypes=[("RAR压缩包", "*.rar"), ("压缩包", "*.rar *.zip"), ("所有文件", "*.*")],
        )
        if not src:
            return
        name = self._ask_replace_novel(novels)
        if not name:
            return
        target = os.path.join(self.root_dir, name)
        if not messagebox.askyesno(
                APP_NAME,
                "将用 RAR 内容替换小说「%s」\n删除原目录：%s\n\n是否继续？" % (name, target)):
            return
        if not self._extract_rar_replace(src, target):
            return
        # 导入后回到主列表：无论当前在小说内部还是列表页，
        # 统一重建列表视图，清掉可能指向已销毁工作区的 current_view
        self.current_novel = None
        self.show_novel_list()
        self._novel_names = self.find_novel_names()
        self.refresh_novel_list()
        messagebox.showinfo(APP_NAME, "RAR 导入完成：\n%s" % target)

    def _ask_replace_novel(self, novels):
        """弹出对话框，让用户选择要替换的现有小说，返回小说名。"""
        dlg = tk.Toplevel(self)
        dlg.title("选择要替换的小说")
        dlg.transient(self)
        dlg.grab_set()
        dlg.geometry("420x200")
        dlg.resizable(False, False)
        ttk.Label(dlg, text="选择要用 RAR 内容替换的现有小说：",
                  padding=(14, 14, 14, 4)).pack(anchor="w")
        var = tk.StringVar(value=novels[0])
        combo = ttk.Combobox(dlg, textvariable=var, values=novels,
                             state="readonly", width=36)
        combo.pack(anchor="w", padx=14, pady=4)
        result = {"name": None}
        def on_ok():
            if var.get():
                result["name"] = var.get()
                dlg.destroy()
        btns = ttk.Frame(dlg)
        btns.pack(side="bottom", fill="x", padx=14, pady=12)
        ttk.Button(btns, text="取消", command=dlg.destroy).pack(side="right", padx=(6, 0))
        ttk.Button(btns, text="替换", command=on_ok).pack(side="right")
        dlg.update_idletasks()
        self.wait_window(dlg)
        return result["name"]

    @staticmethod
    def _kind_of(path):
        low = os.path.basename(path).lower()
        return "7z" if "7z" in low else "unrar"

    def _find_unrar_tool(self):
        """查找本机可用的 RAR 解压工具，返回 (exe路径, 类型)。类型为 'unrar' 或 '7z'。"""
        saved = self.cfg.get("unrar_path") or self._unrar_path
        if saved and os.path.isfile(saved):
            return saved, self._kind_of(saved)
        order = ["UnRAR", "Rar", "WinRAR", "7z", "7za", "7zr"]
        for name in order:
            exe = shutil.which(name)
            if exe:
                return exe, self._kind_of(exe)
        for p in (r"C:\Program Files\WinRAR\UnRAR.exe",
                  r"C:\Program Files\WinRAR\WinRAR.exe",
                  r"C:\Program Files (x86)\WinRAR\UnRAR.exe",
                  r"C:\Program Files (x86)\WinRAR\WinRAR.exe",
                  r"C:\Program Files\7-Zip\7z.exe",
                  r"C:\Program Files (x86)\7-Zip\7z.exe"):
            if os.path.exists(p):
                return p, self._kind_of(p)
        return None, None

    def _ask_unrar_path(self):
        """弹窗让用户输入/选择解压软件的安装路径，校验后保存到配置，返回 (exe, kind)。"""
        while True:
            dlg = tk.Toplevel(self)
            dlg.title("设置 RAR 解压软件")
            dlg.transient(self)
            dlg.grab_set()
            dlg.geometry("520x170")
            dlg.resizable(False, False)
            ttk.Label(dlg, text="未找到 RAR 解压软件，请指定安装路径\n"
                                "（例如 UnRAR.exe 或 WinRAR.exe / 7z.exe）：",
                      padding=(14, 12, 14, 6)).pack(anchor="w")
            var = tk.StringVar()
            entry_row = ttk.Frame(dlg)
            entry_row.pack(fill="x", padx=14)
            ttk.Entry(entry_row, textvariable=var).pack(side="left", fill="x", expand=True)
            def browse():
                p = filedialog.askopenfilename(title="选择解压软件",
                                               filetypes=[("可执行文件", "*.exe"), ("所有文件", "*.*")])
                if p:
                    var.set(p)
            ttk.Button(entry_row, text="...", width=3, command=browse).pack(side="left", padx=(4, 0))
            result = {"path": None, "cancel": True}
            def on_ok():
                p = var.get().strip().strip('"')
                if not p or not os.path.isfile(p):
                    messagebox.showwarning(APP_NAME, "文件不存在，请重新选择。", parent=dlg)
                    return
                result["path"] = p
                result["cancel"] = False
                dlg.destroy()
            def on_cancel():
                dlg.destroy()
            btns = ttk.Frame(dlg)
            btns.pack(side="bottom", fill="x", padx=14, pady=(0, 12))
            ttk.Button(btns, text="取消", command=on_cancel).pack(side="right", padx=(6, 0))
            ttk.Button(btns, text="确定", command=on_ok).pack(side="right")
            dlg.update_idletasks()
            dlg.wait_window(dlg)
            if result["cancel"]:
                return None, None
            # 保存到配置，下次自动识别
            self.cfg["unrar_path"] = result["path"]
            self._unrar_path = result["path"]
            save_config(self.cfg)
            return result["path"], self._kind_of(result["path"])

    @staticmethod
    def _can_create_rar(path):
        base = os.path.basename(path).lower()
        return base in ("winrar.exe", "rar.exe") or base.startswith("winrar")

    def _find_rar_creator(self):
        """查找可创建 RAR 的工具（仅 WinRAR.exe / Rar.exe），返回 exe 路径或 None。"""
        saved = self.cfg.get("unrar_path") or self._unrar_path
        if saved and os.path.isfile(saved) and self._can_create_rar(saved):
            return saved
        order = ["Rar", "WinRAR"]
        for name in order:
            exe = shutil.which(name)
            if exe and self._can_create_rar(exe):
                return exe
        for p in (r"C:\Program Files\WinRAR\WinRAR.exe",
                  r"C:\Program Files\WinRAR\Rar.exe",
                  r"C:\Program Files (x86)\WinRAR\WinRAR.exe",
                  r"C:\Program Files (x86)\WinRAR\Rar.exe"):
            if os.path.exists(p) and self._can_create_rar(p):
                return p
        return None

    def _ask_rar_creator(self):
        """弹窗让用户指定 WinRAR.exe 或 Rar.exe 路径，校验后保存，返回 exe 或 None。"""
        while True:
            dlg = tk.Toplevel(self)
            dlg.title("设置 RAR 压缩软件")
            dlg.transient(self)
            dlg.grab_set()
            dlg.geometry("520x170")
            dlg.resizable(False, False)
            ttk.Label(dlg, text="导出 RAR 需要 WinRAR（7-Zip 无法创建 RAR）\n"
                                "请指定 WinRAR.exe 或 Rar.exe 的安装路径：",
                      padding=(14, 12, 14, 6)).pack(anchor="w")
            var = tk.StringVar()
            entry_row = ttk.Frame(dlg)
            entry_row.pack(fill="x", padx=14)
            ttk.Entry(entry_row, textvariable=var).pack(side="left", fill="x", expand=True)
            def browse():
                p = filedialog.askopenfilename(title="选择压缩软件",
                                               filetypes=[("可执行文件", "*.exe"), ("所有文件", "*.*")])
                if p:
                    var.set(p)
            ttk.Button(entry_row, text="...", width=3, command=browse).pack(side="left", padx=(4, 0))
            result = {"path": None, "cancel": True}
            def on_ok():
                p = var.get().strip().strip('"')
                if not p or not os.path.isfile(p):
                    messagebox.showwarning(APP_NAME, "文件不存在，请重新选择。", parent=dlg)
                    return
                if not self._can_create_rar(p):
                    messagebox.showwarning(APP_NAME, "所选程序无法创建 RAR。\n请选择 WinRAR.exe 或 Rar.exe。",
                                           parent=dlg)
                    return
                result["path"] = p
                result["cancel"] = False
                dlg.destroy()
            def on_cancel():
                dlg.destroy()
            btns = ttk.Frame(dlg)
            btns.pack(side="bottom", fill="x", padx=14, pady=(0, 12))
            ttk.Button(btns, text="取消", command=on_cancel).pack(side="right", padx=(6, 0))
            ttk.Button(btns, text="确定", command=on_ok).pack(side="right")
            dlg.update_idletasks()
            dlg.wait_window(dlg)
            if result["cancel"]:
                return None
            self.cfg["unrar_path"] = result["path"]
            self._unrar_path = result["path"]
            save_config(self.cfg)
            return result["path"]

    def _extract_rar_replace(self, src, target):
        """解压 rar 到临时目录，删除原小说目录，再用解压内容替换。"""
        exe, kind = self._find_unrar_tool()
        if not exe:
            exe, kind = self._ask_unrar_path()
            if not exe:
                messagebox.showinfo(APP_NAME, "已取消导入。")
                return False
        tmp = tempfile.mkdtemp(prefix="ttl_rar_")
        try:
            try:
                if kind == "7z":
                    rc = subprocess.run([exe, "x", src, "-o" + tmp, "-y", "-bso0", "-bsp0"],
                                        capture_output=True, timeout=300).returncode
                else:
                    rc = subprocess.run([exe, "x", "-o+", "-y", src, tmp + os.sep],
                                        capture_output=True, timeout=300).returncode
            except Exception as e:
                messagebox.showerror(APP_NAME, "解压失败：%s" % e)
                return False
            if rc != 0:
                messagebox.showerror(APP_NAME, "解压失败（错误码 %s）。\n请确认文件是有效的 RAR 压缩包。" % rc)
                return False
            content = self._locate_extracted_root(tmp)
            if not content or not os.path.isdir(content):
                messagebox.showerror(APP_NAME, "压缩包内未找到小说内容。")
                return False
            # 删除原小说目录并被替换的内容覆盖
            if os.path.exists(target):
                shutil.rmtree(target)
            ensure_dir(target)
            for item in os.listdir(content):
                s = os.path.join(content, item)
                d = os.path.join(target, item)
                if os.path.isdir(s):
                    shutil.copytree(s, d, dirs_exist_ok=True)
                else:
                    shutil.copy2(s, d)
            return True
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    @staticmethod
    def _locate_extracted_root(tmp):
        """定位解压后的小说根目录：若顶层只有一个目录则返回该目录，否则返回临时目录。"""
        items = [i for i in os.listdir(tmp) if os.path.isdir(os.path.join(tmp, i))]
        if len(items) == 1:
            return os.path.join(tmp, items[0])
        return tmp

    # ================= 导出 =================
    def export_novel_dialog(self):
        if not self.root_dir or not os.path.isdir(self.root_dir):
            if not self._prompt_root_dir():
                return
        if not self._novel_names:
            messagebox.showinfo(APP_NAME, "当前根目录下没有小说。")
            return

        dlg = tk.Toplevel(self)
        dlg.title("导出小说为RAR文件")
        dlg.transient(self)
        dlg.grab_set()
        dlg.geometry("440x300")
        dlg.resizable(False, False)

        ttk.Label(dlg, text="选择要导出的小说：").pack(anchor="w", padx=14, pady=(12, 4))
        var = tk.StringVar(value=self._novel_names[0] if self._novel_names else "")
        combo = ttk.Combobox(dlg, textvariable=var, values=self._novel_names, state="readonly", width=40)
        combo.pack(anchor="w", padx=14)

        ttk.Label(dlg, text="导出到的目录（生成 小说名.rar）：").pack(anchor="w", padx=14, pady=(10, 4))
        dest_frame = ttk.Frame(dlg)
        dest_frame.pack(fill="x", padx=14)
        dest_var = tk.StringVar()
        ttk.Entry(dest_frame, textvariable=dest_var, width=28).pack(side="left", fill="x", expand=True)
        def pick_dest():
            d = filedialog.askdirectory(title="选择导出目录")
            if d:
                dest_var.set(d)
        ttk.Button(dest_frame, text="...", width=3, command=pick_dest).pack(side="left", padx=(4, 0))

        exclude_backup = tk.BooleanVar(value=True)
        ttk.Checkbutton(dlg, text="导出时排除 _备份 目录", variable=exclude_backup).pack(anchor="w", padx=14, pady=(12, 0))

        def on_export():
            name = var.get()
            dest = dest_var.get()
            if not name:
                return
            if not dest:
                messagebox.showwarning("导出", "请选择导出目录。", parent=dlg)
                return
            dlg.destroy()
            self._do_export(name, dest, exclude_backup.get())

        btns = ttk.Frame(dlg)
        btns.pack(side="bottom", fill="x", padx=14, pady=12)
        ttk.Button(btns, text="取消", command=dlg.destroy).pack(side="right", padx=(6, 0))
        ttk.Button(btns, text="导出", command=on_export).pack(side="right")

        dlg.update_idletasks()

    def _do_export(self, name, dest, exclude_backup):
        src = os.path.join(self.root_dir, name)
        rar_file = os.path.join(dest, name + ".rar")
        creator = self._find_rar_creator()
        if not creator:
            creator = self._ask_rar_creator()
            if not creator:
                messagebox.showinfo(APP_NAME, "已取消导出。")
                return
        if os.path.exists(rar_file):
            os.remove(rar_file)
        # 排除 _备份：暂时移出同目录，压缩完再移回
        moved = None
        backup = os.path.join(src, SUBDIR_BACKUP)
        if exclude_backup and os.path.isdir(backup):
            moved = backup + ".ttltmp"
            if os.path.exists(moved):
                shutil.rmtree(moved, ignore_errors=True)
            shutil.move(backup, moved)
        try:
            rc = subprocess.run([creator, "a", "-r", "-ep1", "-y",
                                 rar_file, src],
                                capture_output=True, timeout=600).returncode
        except Exception as e:
            messagebox.showerror(APP_NAME, "导出失败：%s" % e)
            return
        finally:
            if moved and os.path.exists(moved):
                try:
                    shutil.move(moved, backup)
                except Exception:
                    pass
        if rc != 0 or not os.path.exists(rar_file):
            messagebox.showerror(APP_NAME, "导出失败（错误码 %s）。%s" % (rc, (creator and "请检查软件路径。") or ""))
            return
        messagebox.showinfo(APP_NAME, "导出完成：\n%s" % rar_file)

    # ================= 关于/关闭 =================
    def show_about(self):
        about = ("%s v%s\n\n" % (APP_NAME, APP_VERSION) +
                 "一个简单易用的小说编写管理器。\n"
                 "支持草稿、正文、角色库、世界观、章节管理，\n"
                 "RAR 导入覆盖小说目录，并将小说导出为 RAR 文件。")
        messagebox.showinfo("关于 " + APP_NAME, about)

    def _on_close(self):
        if self.current_view is not None:
            try:
                sv = getattr(self.current_view, "save_current", None)
                if sv:
                    sv()
            except Exception:
                pass
        self.destroy()


# ---------------------------------------------------------------------------
# 小说列表视图（主界面）
# ---------------------------------------------------------------------------
class NovelListView(ttk.Frame):
    """主界面：显示小说列表，每项为 封面图片 + 小说名。点击进入内容管理界面。"""

    CARD_W = 160
    CARD_H = 210

    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.selected = None
        self.photo_cache = {}
        self.card_frames = {}
        self._build_ui()
        self.refresh()

    # ---------------- UI ----------------
    def _build_ui(self):
        header = ttk.Frame(self, padding=(12, 8))
        header.pack(fill="x")
        ttk.Label(header, text="我的小说",
                  font=("Microsoft YaHei UI", 15, "bold")).pack(side="left")

        right = ttk.Frame(header)
        right.pack(side="right")
        ttk.Button(right, text="设置路径", command=self.app.set_root_dir).pack(side="left", padx=3)
        ttk.Button(right, text="新建小说", command=self.app.new_novel).pack(side="left", padx=3)
        ttk.Button(right, text="导入RAR", command=self.app.import_rar).pack(side="left", padx=3)
        ttk.Button(right, text="导出", command=self.app.export_novel_dialog).pack(side="left", padx=3)
        ttk.Button(right, text="刷新", command=self.app.refresh_novel_list).pack(side="left", padx=3)

        # 提示条
        ttk.Label(self, text="双击小说封面或名称进入内容管理",
                  foreground="#888").pack(anchor="w", padx=12)

        # 滚动容器
        wrap = ttk.Frame(self)
        wrap.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        self.canvas = tk.Canvas(wrap, highlightthickness=0, bg="#f7f7f7")
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.grid_frame = ttk.Frame(self.canvas)
        self.grid_window = self.canvas.create_window((8, 8), window=self.grid_frame, anchor="nw")
        self.grid_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<Enter>", lambda e: self.canvas.bind_all("<MouseWheel>",
                           lambda ev: self.canvas.yview_scroll(int(-ev.delta / 120), "units")))
        self.canvas.bind("<Leave>", lambda e: self.canvas.unbind_all("<MouseWheel>"))

        self.placeholder = ttk.Label(
            self.grid_frame,
            text="尚未设置小说根目录。\n\n点击右上角「设置路径」选择存放小说的目录，\n或使用菜单「文件 → 设置小说根目录」。",
            font=("Microsoft YaHei UI", 12), foreground="#999", justify="center")
        self.placeholder.grid(row=0, column=0, pady=120)

    def _on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # 根据可用宽度重排卡片列数
        width = event.width
        cols = max(1, (width - 8) // (self.CARD_W + 16))
        self._cols = cols
        self._reflow()

    # ---------------- 渲染 ----------------
    def refresh(self):
        self.app._novel_names = self.app.find_novel_names()
        names = self.app._novel_names
        for w in self.grid_frame.winfo_children():
            w.destroy()
        self.photo_cache.clear()
        self.card_frames.clear()

        if not names:
            # 重新创建占位提示（因为上面已销毁所有子控件）
            self.placeholder = ttk.Label(
                self.grid_frame,
                text="尚未设置小说根目录。\n\n点击右上角「设置路径」选择存放小说的目录，\n或使用菜单「文件 → 设置小说根目录」。",
                font=("Microsoft YaHei UI", 12), foreground="#999", justify="center")
            self.placeholder.grid(row=0, column=0, pady=120)
            return
        self._reflow()

    def _build_cards(self):
        names = self.app._novel_names
        for i, name in enumerate(names):
            novel_dir = os.path.join(self.app.root_dir, name)
            card = self._make_card(name, novel_dir)
            self.card_frames[i] = card

    def _make_card(self, name, novel_dir):
        card = tk.Frame(self.grid_frame, width=self.CARD_W, height=self.CARD_H,
                        bg="white", highlightbackground="#ddd", highlightthickness=1,
                        cursor="hand2")
        card.pack_propagate(False)

        img = self.app._load_cover_photo(novel_dir, size=120)
        if img:
            photo = img
        else:
            photo = None
        self.photo_cache[name] = photo

        cover_box = tk.Label(card, bg="white", image=photo, text="" if photo else "无封面",
                             width=self.CARD_W, height=130, compound="center")
        cover_box.pack(padx=6, pady=(8, 4))

        title = tk.Label(card, text=name, bg="white", font=("Microsoft YaHei UI", 11, "bold"),
                         wraplength=self.CARD_W - 12, justify="center")
        title.pack(side="bottom", pady=(0, 8), fill="x")

        for w in (card, cover_box, title):
            w.bind("<Button-1>", lambda e, n=name, d=novel_dir: self._on_click(n, d))
            w.bind("<Double-Button-1>", lambda e, n=name, d=novel_dir: self.app.show_workspace(d))

        return card

    def _reflow(self):
        for w in self.grid_frame.winfo_children():
            if w.winfo_manager() != "":
                w.grid_forget()
        self._build_cards()
        cols = getattr(self, "_cols", 4)
        for i, name in enumerate(self.app._novel_names):
            cl = i % cols
            rw = i // cols
            card = self.card_frames.get(i)
            if card:
                card.grid(row=rw, column=cl, padx=8, pady=8, sticky="n")

    def _on_click(self, name, novel_dir):
        # 单击选中
        self.selected = name
        for i, other in enumerate(self.app._novel_names):
            other_dir = os.path.join(self.app.root_dir, other)
            card = self.card_frames.get(i)
            if card:
                hl = (other == name)
                card.configure(highlightbackground="#1f6feb" if hl else "#ddd",
                               highlightthickness=3 if hl else 1)

    def select_by_name(self, name):
        self.selected = name
        if name in self.app._novel_names:
            self._on_click(name, os.path.join(self.app.root_dir, name))

    def save_current(self):
        pass


# ---------------------------------------------------------------------------
# 小说工作区（编辑器）
# ---------------------------------------------------------------------------
class NovelWorkspace(ttk.Frame):
    def __init__(self, master, app, novel_dir):
        super().__init__(master)
        self.app = app
        self.novel_dir = novel_dir
        self.novel_name = os.path.basename(novel_dir)

        self.current_path = None
        self.saved_content_marker = None
        self.current_section = None
        self.item_map = {}
        self.auto_save_after = None

        self._build_ui()
        self.load_section(SUBDIR_BODY)

    # ---------------- UI 构建 ----------------
    def _build_ui(self):
        # 顶部标题条
        title = ttk.Frame(self, padding=(10, 8))
        title.pack(fill="x")
        # 封面平铺区：展示封面目录下所有图片（横向滚动）
        cover_strip_outer = ttk.Frame(title)
        cover_strip_outer.pack(side="left", padx=(0, 10), fill="y")
        self.cover_strip = tk.Canvas(cover_strip_outer, height=66, width=340,
                                      highlightthickness=0, bg="#eef0f2")
        cov_hsb = ttk.Scrollbar(cover_strip_outer, orient="horizontal",
                                command=self.cover_strip.xview)
        self.cover_strip.configure(xscrollcommand=cov_hsb.set)
        self.cover_strip.pack(side="top", fill="x")
        cov_hsb.pack(side="bottom", fill="x")
        self.cover_imgs = []  # 保留引用防止被 GC
        self._refresh_cover_strip()
        ttk.Label(title, text="小说：" + self.novel_name,
                  font=("Microsoft YaHei UI", 13, "bold")).pack(side="left")

        tb = ttk.Frame(title)
        tb.pack(side="right")
        ttk.Button(tb, text="上传封面", command=self.app.upload_cover).pack(side="left", padx=2)
        ttk.Button(tb, text="导入RAR", command=self.app.import_rar).pack(side="left", padx=2)
        ttk.Button(tb, text="导出", command=self.app.export_novel_dialog).pack(side="left", padx=2)
        ttk.Button(tb, text="返回", command=self._back_to_list).pack(side="left", padx=2)

        # 主体：左(文件夹) + 中(文件列表) + 右(编辑器)
        main = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        main.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # 左：主要文件夹
        left = ttk.Frame(main, width=130)
        left.pack_propagate(False)
        ttk.Label(left, text="文件夹", font=("Microsoft YaHei UI", 9, "bold"),
                  foreground="#666").pack(anchor="w", padx=6, pady=(4, 2))
        self.folder_tree = ttk.Treeview(left, show="tree", selectmode="browse")
        self.folder_tree.pack(fill="both", expand=True, padx=(2, 2), pady=(0, 4))
        self.folder_tree.bind("<<TreeviewSelect>>", self._on_folder_select)

        # 中：文件列表（当前文件夹下的文件）+ 筛选 + 滚动条
        middle = ttk.Frame(main, width=270)
        middle.pack_propagate(False)
        filter_row = ttk.Frame(middle)
        filter_row.pack(fill="x", padx=6, pady=(4, 2))
        ttk.Label(filter_row, text="筛选:").pack(side="left")
        self.filter_var = tk.StringVar()
        self.filter_entry = ttk.Entry(filter_row, textvariable=self.filter_var)
        self.filter_entry.pack(side="left", fill="x", expand=True, padx=(4, 0))
        self.filter_var.trace_add("write", self._on_filter_change)
        self.file_header = ttk.Label(middle, text="正文", font=("Microsoft YaHei UI", 9, "bold"),
                                     foreground="#666")
        self.file_header.pack(anchor="w", padx=6, pady=(0, 2))
        list_frame = ttk.Frame(middle)
        list_frame.pack(fill="both", expand=True, padx=(2, 2), pady=(0, 4))
        self.file_tree = ttk.Treeview(list_frame, show="tree", selectmode="browse")
        self.file_vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=self.file_vsb.set)
        self.file_tree.pack(side="left", fill="both", expand=True)
        self.file_vsb.pack(side="right", fill="y")
        self.file_tree.bind("<<TreeviewSelect>>", self._on_file_select)

        # 右：编辑器 / 预览（按需切换）
        right = ttk.Frame(main)
        main.add(left, weight=1)
        main.add(middle, weight=1)
        main.add(right, weight=4)

        # 编辑器容器
        self.editor_frame = ttk.Frame(right)
        # 编辑器工具栏
        edit_bar = ttk.Frame(self.editor_frame)
        edit_bar.pack(fill="x", pady=(0, 4))
        self.bar_title = ttk.Label(edit_bar, text="", font=("Microsoft YaHei UI", 11, "bold"))
        self.bar_title.pack(side="left")
        self.word_label = ttk.Label(edit_bar, text=" 字数：0")
        self.word_label.pack(side="right")
        self.auto_save = tk.BooleanVar(value=False)
        ttk.Checkbutton(edit_bar, text="自动保存", variable=self.auto_save).pack(side="right", padx=4)
        ttk.Button(edit_bar, text="保存", command=self.save_current).pack(side="right", padx=2)
        ttk.Button(edit_bar, text="重命名", command=self.rename_item).pack(side="right", padx=2)
        ttk.Button(edit_bar, text="删除", command=self.delete_item).pack(side="right", padx=2)
        ttk.Button(edit_bar, text="新建", command=self.new_item).pack(side="right", padx=2)

        # 编辑区
        self.text = scrolledtext.ScrolledText(self.editor_frame, wrap=tk.WORD,
                                              font=("Microsoft YaHei UI", 11),
                                              undo=True, padx=8, pady=8)
        self.text.pack(fill="both", expand=True)
        self.text.bind("<KeyRelease>", self._on_text_change)

        # 预览容器
        self.preview_frame = ttk.Frame(right)
        self.preview_content = ttk.Frame(self.preview_frame)
        self.preview_content.pack(fill="both", expand=True)
        self.preview_open_btn = ttk.Button(self.preview_frame, text="用系统程序打开",
                                            command=self._open_current_external)
        self.preview_open_btn.pack(side="bottom", anchor="e", padx=8, pady=(2, 6))

        # 默认显示编辑器
        self._active_panel = "editor"
        self.editor_frame.pack(fill="both", expand=True)
        self.preview_frame.pack_forget()
        self._preview_img_path = None
        self._preview_photo = None

        # 状态栏
        self.status_bar = ttk.Label(self, text="就绪", anchor="w", relief=tk.SUNKEN, padding=(6, 2))
        self.status_bar.pack(fill="x", side="bottom")

        self._init_folders()

    def _refresh_cover_strip(self):
        """在顶部平铺展示封面目录下所有图片。"""
        try:
            self.cover_imgs = []
            self.cover_strip.delete("all")
            x = 3
            cover_dir = os.path.join(self.novel_dir, SUBDIR_COVER)
            if os.path.isdir(cover_dir):
                names = sorted([n for n in os.listdir(cover_dir)
                                if os.path.isfile(os.path.join(cover_dir, n))
                                and is_image_file(n)])
                for n in names:
                    try:
                        img = Image.open(os.path.join(cover_dir, n))
                        img.thumbnail((60, 60), Image.LANCZOS)
                        photo = ImageTk.PhotoImage(img)
                        self.cover_imgs.append(photo)
                        self.cover_strip.create_image(x, 3, anchor="nw", image=photo)
                        x += 66
                    except Exception:
                        continue
                if names:
                    self.cover_strip.configure(scrollregion=(0, 0, max(x, 340), 66))
                    return
            # 无封面
            self.cover_strip.configure(scrollregion=(0, 0, 340, 66))
            self.cover_strip.create_text(6, 33, anchor="w", text="无封面", fill="#999")
        except Exception:
            pass

    def refresh_cover(self):
        self._refresh_cover_strip()

    def _back_to_list(self):
        if self.auto_save.get():
            self.save_current()
        self.app.current_novel = None
        self.app.show_novel_list()

    # ---------------- 文件夹列表(左) + 文件列表(中) ----------------
    def _init_folders(self):
        """左栏文件夹定义：(显示名, 定位键, 是否为目录文件夹)。"""
        self.folders = [
            ("正文", SUBDIR_BODY, True),
            ("封面", SUBDIR_COVER, True),
            ("草稿", SUBDIR_DRAFTS, True),
            ("角色库", SUBDIR_ROLES, True),
            ("世界观", FILE_WORLDVIEW, False),
            ("其他文件", "__other__", True),
        ]
        self.folder_tree.delete(*self.folder_tree.get_children())
        self._folder_item_index = {}
        for i, (display, _, _) in enumerate(self.folders):
            item = self.folder_tree.insert("", "end", text=display)
            self._folder_item_index[item] = i
        # 默认选中“正文”（Treeview 自带选中底色）
        self._active_folder_index = 0
        first = self.folder_tree.get_children()[0] if self.folder_tree.get_children() else None
        if first:
            self.folder_tree.selection_set(first)
            self.folder_tree.focus(first)
        self._load_files_for_index(0)

    def _folder_full_path(self, folder):
        """根据文件夹项返回其对应的主目录路径(目录文件夹)或文件路径(单文件)。"""
        key = folder
        if key == SUBDIR_BODY:
            return os.path.join(self.novel_dir, SUBDIR_BODY), True
        if key == SUBDIR_DRAFTS:
            return os.path.join(self.novel_dir, SUBDIR_DRAFTS), True
        if key == SUBDIR_ROLES:
            return os.path.join(self.novel_dir, SUBDIR_ROLES), True
        if key == SUBDIR_OTHER:
            return self.novel_dir, True
        # 单文件类
        return os.path.join(self.novel_dir, key), False

    def _chapter_title_map(self):
        """解析 章节标题列表.md，返回 {章节数字: 标题} 映射。
        支持格式：
          - 章节序号||标题        例：2||凡人接剑
          - chapter_编号 —— 标题   例：chapter_003 —— 风起
          - 第X章 标题           例：第一章 风起
        """
        path = os.path.join(self.novel_dir, FILE_CHAPTER_TITLES)
        content = read_text(path)
        result = {}

        def cn_to_num(s):
            s = s.strip()
            if s.isdigit():
                return int(s)
            cn = {"零": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4,
                  "五": 5, "六": 6, "七": 7, "八": 8, "九": 9,
                  "十": 10, "百": 100, "千": 1000}
            total, n = 0, 0
            for ch in s:
                if ch in cn:
                    d = cn[ch]
                    if d >= 10:
                        total += (n if n else 1) * d
                        n = 0
                    else:
                        n = d
                else:
                    n = 0
            return total + n

        for raw in content.splitlines():
            line = raw.lstrip(" \t\u3000-#*•·、")
            num = None
            tail = ""

            # 格式1：章节序号||标题
            m = re.match(r"(\d+)\s*[|丨\\｜]\s*(.*)", line)
            if m:
                num = str(int(m.group(1)))
                tail = m.group(2)
            else:
                # 格式2：chapter_编号
                m = re.search(r"chapter[_\s]*(\d+)", line, re.IGNORECASE)
                if m:
                    num = m.group(1)
                    tail = line[m.end():]
                else:
                    # 格式3：第X章
                    m = re.search(r"第\s*(?:(\d+)|([零一二两三四五六七八九十百千]+))\s*章\s*(.*)", line)
                    if m:
                        if m.group(1) is not None:
                            num = str(int(m.group(1)))
                        else:
                            num = str(cn_to_num(m.group(2)) or 0)
                        tail = m.group(3)
            if num is None:
                continue

            t = re.sub(r"^[\s\-–—－—·|:：,，、.。\u3000]+", "", tail)
            t = re.sub(r"[\s\-–—－—·|:：,，、.。\u3000]+$", "", t)
            if t.startswith("（") and t.endswith("）"):
                t = t[1:-1]
            elif t.startswith("(") and t.endswith(")"):
                t = t[1:-1]
            if t:
                result[num.lstrip("0") or num] = t
        return result

    def _file_display(self, key, fname, title_map=None):
        """生成文件列表的显示文本：正文章节附加章节名。"""
        if key == SUBDIR_BODY:
            m = re.match(r"chapter[_\s]*(\d+)", fname, re.IGNORECASE)
            if m:
                num = m.group(1).lstrip("0") or m.group(1)
                title = (title_map or {}).get(num)
                if title:
                    return "%s  ——  %s" % (fname, title)
        return fname

    def _set_folder_selection(self, idx):
        for item, i in self._folder_item_index.items():
            if i == idx:
                self.folder_tree.selection_set(item)
                self.folder_tree.focus(item)
                self.folder_tree.see(item)
                return

    def _load_files_for_index(self, idx):
        if idx < 0 or idx >= len(self.folders):
            return
        self._active_folder_index = idx
        self._populate_file_list()
        # 自动选中并打开第一个文件
        if self._current_folder_files:
            self._open_file(self._current_folder_files[0])
        else:
            self._open_file(None)

    def _populate_file_list(self, select_path=None):
        """根据当前文件夹与筛选词重建文件列表，可指定要重新选中的文件。"""
        idx = getattr(self, "_active_folder_index", 0)
        if idx < 0 or idx >= len(self.folders):
            idx = 0
        self._active_folder_index = idx
        display, key, is_dir_folder = self.folders[idx]
        keep_open = select_path if select_path else self.current_path

        keyword = (self.filter_var.get() or "").strip()
        self.file_header.config(text=display)
        self.file_tree.delete(*self.file_tree.get_children())
        self._current_folder_files = []
        self._file_item_path = {}

        if is_dir_folder:
            target_dir, _ = self._folder_full_path(key)
            title_map = self._chapter_title_map() if key == SUBDIR_BODY else {}
            if os.path.isdir(target_dir):
                if key == SUBDIR_OTHER:  # noqa
                    # 其他文件：列出主目录下所有文件（含 md、图片、xlsx 等），
                    # 但世界观.md 已有独立文件夹，不在此重复显示
                    files = [f for f in sorted(os.listdir(target_dir))
                             if os.path.isfile(os.path.join(target_dir, f))
                             and f != FILE_WORLDVIEW]
                elif key == SUBDIR_COVER:  # noqa
                    # 封面：列出所有图片文件
                    files = [f for f in sorted(os.listdir(target_dir))
                             if os.path.isfile(os.path.join(target_dir, f))
                             and is_image_file(f)]
                else:
                    files = [f for f in sorted(os.listdir(target_dir))
                             if os.path.isfile(os.path.join(target_dir, f))
                             and f.endswith((".txt", ".md"))]
                if keyword:
                    files = [f for f in files if keyword.lower() in f.lower()]
                for f in files:
                    item = self.file_tree.insert("", "end",
                                                 text=self._file_display(key, f, title_map))
                    p = os.path.join(target_dir, f)
                    self._current_folder_files.append(p)
                    self._file_item_path[item] = p
        else:
            path, _ = self._folder_full_path(key)
            item = self.file_tree.insert("", "end", text=key)
            self._current_folder_files.append(path)
            self._file_item_path[item] = path

        if not self._current_folder_files:
            msg = display + ("（空）" if not keyword else "（无匹配）")
            self.file_header.config(text=msg)
            return

        # 选中目标或第一个
        chosen = 0
        if keep_open and keep_open in self._current_folder_files:
            chosen = self._current_folder_files.index(keep_open)
        item_id = self.file_tree.get_children()[chosen]
        self.file_tree.selection_set(item_id)
        self.file_tree.focus(item_id)
        self.file_tree.see(item_id)

    def _on_folder_select(self, event):
        sel = self.folder_tree.selection()
        if not sel:
            return
        idx = self._folder_item_index.get(sel[0])
        if idx is None:
            return
        self._load_files_for_index(idx)

    def _on_file_select(self, event):
        sel = self.file_tree.selection()
        if not sel:
            return
        path = self._file_item_path.get(sel[0])
        if path:
            self._open_file(path)

    def _on_filter_change(self, *args):
        """筛选词变化：重建当前文件列表，若当前打开文件被过滤掉则打开首个。"""
        current = self.current_path
        self._populate_file_list()
        if current and current not in self._current_folder_files and self._current_folder_files:
            self._open_file(self._current_folder_files[0])

    def _open_file(self, path):
        if self.auto_save.get():
            self.save_current()
        if path is None:
            self.current_path = None
            self.text.delete("1.0", tk.END)
            self.bar_title.config(text="")
            self.word_label.config(text=" 字数：0")
            return
        self.current_path = path
        self._preview_img_path = None
        ext = os.path.splitext(path)[1].lower()
        self.bar_title.config(text=os.path.basename(path))
        self.status_bar.config(text="打开：%s" % self._relpath(path))
        if is_image_file(path):
            self._show_panel("preview")
            self._show_preview_image(path)
        elif ext in (".xlsx", ".xlsm"):
            self._show_panel("preview")
            self._show_preview_xlsx(path)
        elif ext in (".txt", ".md"):
            self._show_panel("editor")
            content = read_text(path)
            self.text.delete("1.0", tk.END)
            self.text.insert("1.0", content)
            self.saved_content_marker = self._content_sig()
            self._update_word_count()
        else:
            self._show_panel("preview")
            self._show_preview_info(path)

    def _show_panel(self, name):
        """切换右侧面板：'editor' 或 'preview'。"""
        if getattr(self, "_active_panel", None) == name:
            return
        self._active_panel = name
        if name == "preview":
            self.editor_frame.pack_forget()
            self.preview_frame.pack(fill="both", expand=True)
        else:
            self.preview_frame.pack_forget()
            self.editor_frame.pack(fill="both", expand=True)

    def _clear_preview(self):
        for w in self.preview_content.winfo_children():
            w.destroy()

    def _show_preview_image(self, path):
        self._clear_preview()
        self._preview_img_path = path
        canvas = tk.Canvas(self.preview_content, bg="#f5f5f5")
        canvas.pack(fill="both", expand=True)
        self.preview_canvas = canvas
        canvas.bind("<Configure>", lambda e: self._fit_preview_image())
        self._fit_preview_image()

    def _fit_preview_image(self):
        canvas = getattr(self, "preview_canvas", None)
        if not canvas or not self._preview_img_path:
            return
        try:
            img = Image.open(self._preview_img_path)
            w = canvas.winfo_width() or 400
            h = canvas.winfo_height() or 300
            if w < 10:
                w = 400
            if h < 10:
                h = 300
            scale = min(w / img.width, h / img.height, 1.0)
            nw, nh = max(1, int(img.width * scale)), max(1, int(img.height * scale))
            img = img.resize((nw, nh), Image.LANCZOS)
            self._preview_photo = ImageTk.PhotoImage(img)
            canvas.delete("all")
            canvas.create_image(w // 2, h // 2, anchor="center", image=self._preview_photo)
        except Exception:
            pass

    def _show_preview_xlsx(self, path):
        self._clear_preview()
        try:
            rows, total_rows, total_cols, merged, covered = _read_xlsx_with_styles(path)
        except Exception as e:
            self._show_preview_info(path, "无法解析 xlsx：%s" % e)
            return
        if not rows:
            ttk.Label(self.preview_content, text="（空表格）", padding=(12, 12)).pack(anchor="nw")
            return
        maxcol = max(len(r) for r in rows) or 1
        # 超出预览上限时给出提示
        if total_rows > len(rows) or total_cols > maxcol:
            ttk.Label(self.preview_content,
                      text="仅预览前 %d 行 / %d 列（原表约 %d 行 × %d 列）"
                           % (len(rows), maxcol, total_rows, total_cols),
                      font=("Microsoft YaHei UI", 9), foreground="#666666").pack(anchor="nw")
        # 滚动画布 + 内部网格（每个单元格为带背景色的 Label）
        canvas = tk.Canvas(self.preview_content, bg="white")
        vbar = ttk.Scrollbar(self.preview_content, orient="vertical", command=canvas.yview)
        hbar = ttk.Scrollbar(self.preview_content, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)
        inner = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        for ri, row in enumerate(rows):
            for ci in range(maxcol):
                if (ri, ci) in covered:
                    continue
                cell = row[ci] if ci < len(row) else {}
                val = cell.get("v", "") if isinstance(cell, dict) else cell
                bg = cell.get("bg") if isinstance(cell, dict) else None
                if not bg:
                    bg = "#ffffff" if ri else "#e8eef7"
                is_header = (ri == 0)
                anchor = cell.get("anchor", "w") if isinstance(cell, dict) else "w"
                just = "center" if anchor == "center" else ("right" if anchor == "e" else "left")
                # 优先使用单元格自身字体颜色，否则按背景亮度自动选择
                fg = cell.get("fg") if isinstance(cell, dict) else None
                if not fg:
                    try:
                        r = int(bg[1:3], 16); g = int(bg[3:5], 16); b = int(bg[5:7], 16)
                        lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
                        fg = "#ffffff" if lum < 0.55 else "#000000"
                    except Exception:
                        fg = "#000000"
                span = merged.get((ri, ci))
                bordered = cell.get("bordered", False)
                bcolor = cell.get("border_color") or "#cccccc"
                lbl_kw = dict(text=str(val), bg=bg, fg=fg,
                              font=("Microsoft YaHei UI", 10,
                                    "bold" if is_header else "normal"),
                              width=18, anchor=anchor, padx=3, pady=1,
                              wraplength=160, justify=just,
                              relief="flat", borderwidth=0)
                if bordered:
                    lbl_kw["highlightbackground"] = bcolor
                    lbl_kw["highlightthickness"] = 1
                else:
                    lbl_kw["highlightthickness"] = 0
                lbl = tk.Label(inner, **lbl_kw)
                grid_kw = {"row": ri, "column": ci, "sticky": "nsew"}
                if span:
                    grid_kw["rowspan"] = span[0]
                    grid_kw["columnspan"] = span[1]
                lbl.grid(**grid_kw)
        inner.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
        vbar.pack(side="right", fill="y")
        hbar.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)
        self.preview_canvas = canvas

    def _show_preview_info(self, path, msg=None):
        self._clear_preview()
        try:
            size = os.path.getsize(path)
        except Exception:
            size = 0
        info = (msg or ("该文件类型暂不支持预览：%s" % os.path.splitext(path)[1])) + \
               "\n\n文件名：%s\n大小：%d 字节" % (os.path.basename(path), size)
        ttk.Label(self.preview_content, text=info, justify="left",
                  padding=(12, 12), font=("Microsoft YaHei UI", 11)).pack(anchor="nw")

    def _open_current_external(self):
        if not self.current_path or not os.path.exists(self.current_path):
            return
        try:
            os.startfile(self.current_path)
        except Exception as e:
            messagebox.showerror(APP_NAME, "打开失败：%s" % e)

    def load_section(self, section):
        """加载某个分类目录下的文件（默认正文）。"""
        for idx, (display, key, is_dir_folder) in enumerate(self.folders):
            if key == section:
                self._set_folder_selection(idx)
                self._load_files_for_index(idx)
                return

    def _relpath(self, path):
        try:
            return os.path.relpath(path, self.novel_dir)
        except Exception:
            return path

    # ---------------- 编辑器操作 ----------------
    def _content_sig(self):
        return (self.text.get("1.0", "end-1c"), self.current_path)

    def _on_text_change(self, event=None):
        self._update_word_count()
        self._schedule_autosave_clear()

    def _schedule_autosave_clear(self):
        # 不自动保存，仅标记标题星号
        if self.current_path and self.saved_content_marker:
            if self._content_sig()[0] != self.saved_content_marker[0]:
                if not self.bar_title.cget("text").endswith(" *"):
                    self.bar_title.config(text=os.path.basename(self.current_path) + " *")

    def _update_word_count(self):
        txt = self.text.get("1.0", "end-1c")
        # 字数：中文按字符，英文按单词
        count = len(txt.replace("\n", "").replace(" ", ""))
        self.word_label.config(text="  字数：%d" % count)

    def save_current(self):
        if not self.current_path:
            return
        if not self.current_path.lower().endswith((".txt", ".md")):
            return
        content = self.text.get("1.0", "end-1c")
        write_text_smart(self.current_path, content)
        self.saved_content_marker = self._content_sig()
        self.bar_title.config(text=os.path.basename(self.current_path))
        self.status_bar.config(text="已保存：%s" % self._relpath(self.current_path))
        self._update_word_count()

    # ---------------- 新建 ----------------
    def new_item(self):
        if not self.current_path:
            return
        cur_dir = os.path.dirname(self.current_path)
        cur_section = self._section_of_path(self.current_path)

        if cur_section == SUBDIR_BODY:
            # 正文：自动下一章编号
            chapter_num = self._next_chapter_number(cur_dir)
            name = "chapter_%03d" % chapter_num
            ext = self._default_ext()
            default_name = name + ext
        elif cur_section == SUBDIR_DRAFTS:
            default_name = draft_default_name(".md")
        elif cur_section == SUBDIR_ROLES:
            default_name = "新角色.md"
        else:
            default_name = "新文件.md"

        name = simpledialog.askstring("新建", "文件名（含扩展名）：", initialvalue=default_name, parent=self)
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if not name.lower().endswith((".txt", ".md")):
            name += ".md"
        target = os.path.join(cur_dir, name)
        if os.path.exists(target):
            messagebox.showwarning(APP_NAME, "同名文件已存在。")
            return
        write_text(target, "# " + os.path.splitext(name)[0] + "\n\n")
        self._reload_current_files(select=target)

    def _default_ext(self):
        return ".md"

    def _next_chapter_number(self, body_dir):
        max_num = 0
        if os.path.isdir(body_dir):
            for f in os.listdir(body_dir):
                base = os.path.splitext(f)[0]
                if base.lower().startswith("chapter_"):
                    num = base[len("chapter_"):]
                    nums = "".join(ch for ch in num if ch.isdigit())
                    if nums:
                        try:
                            max_num = max(max_num, int(nums))
                        except Exception:
                            pass
        return max_num + 1

    def _section_of_path(self, path):
        rel = self._relpath(path)
        parts = rel.split(os.sep)
        if parts:
            return parts[0]
        return ""

    # ---------------- 重命名 ----------------
    def rename_item(self):
        if not self.current_path:
            return
        cur = os.path.basename(self.current_path)
        name = simpledialog.askstring("重命名", "新文件名（含扩展名）：", initialvalue=cur, parent=self)
        if not name or name.strip() == cur:
            return
        name = name.strip()
        if not name:
            return
        if os.path.sep in name or "/" in name:
            messagebox.showwarning(APP_NAME, "文件名不能包含路径分隔符。")
            return
        target = os.path.join(os.path.dirname(self.current_path), name)
        if os.path.exists(target):
            messagebox.showwarning(APP_NAME, "同名文件已存在。")
            return
        os.rename(self.current_path, target)
        self.current_path = target
        self.saved_content_marker = self._content_sig()
        self._reload_current_files(select=target)
        self.bar_title.config(text=os.path.basename(target))
        self.status_bar.config(text="已重命名")

    # ---------------- 删除 ----------------
    def delete_item(self):
        if not self.current_path:
            return
        if not messagebox.askyesno("删除", "确定删除文件：\n%s" % os.path.basename(self.current_path)):
            return
        path = self.current_path
        os.remove(path)
        self.current_path = None
        self.text.delete("1.0", tk.END)
        self.bar_title.config(text="")
        self.status_bar.config(text="已删除")
        self._reload_current_files()

    # ---------------- 刷新 ----------------
    def _reload_current_files(self, select=None):
        """重新加载当前选中文件夹的文件列表，可指定要选中的文件路径。"""
        self._populate_file_list(select_path=select)
        if select and select in self._current_folder_files:
            self._open_file(select)

    def destroy(self):
        try:
            if self.auto_save.get():
                self.save_current()
        except Exception:
            pass
        super().destroy()


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------
def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
